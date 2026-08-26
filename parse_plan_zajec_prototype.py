import openpyxl
import datetime
import json

wb = openpyxl.load_workbook('Rok 3 2025-2026/🦠PLAN ZAJĘĆ🦠/III-rok-25-26.xlsx', data_only=True)
ws = wb['Arkusz2']

MONTHS = {
    'Październik': 10, 'Listopad': 11, 'Grudzień': 12, 'Styczeń': 1,
    'Luty': 2, 'Marzec': 3, 'Kwiecień': 4, 'Maj': 5, 'Czerwiec': 6,
}
# markitdown/openpyxl czasem gubi polskie znaki (encoding), więc dopasowanie po prefiksie
def month_num(label):
    if not label:
        return None
    for name, num in MONTHS.items():
        if label[:3].lower() == name[:3].lower() or label.startswith(name[:2]):
            return num
    return None

# Zbuduj mapowanie kolumna -> data (rok 2025 dla X-XII, 2026 dla I-VI), forward-fill etykiety miesiąca.
col_dates = {}
current_month = None
for c in range(3, 154):
    label = ws.cell(row=60, column=c).value
    if label and label != 'GRUPY':
        m = month_num(label)
        if m:
            current_month = m
    day = ws.cell(row=62, column=c).value
    if day is None or current_month is None:
        continue
    year = 2025 if current_month >= 10 else 2026
    try:
        col_dates[c] = datetime.date(year, current_month, int(day))
    except ValueError:
        pass

# Legenda: kod -> nazwa przedmiotu
legend = {}
for r in range(12, 55):
    code = ws.cell(row=r, column=3).value
    name = ws.cell(row=r, column=4).value
    if code and name and code not in legend:
        legend[code] = name.strip()

CODE_TO_SUBJECT_KEY = {
    'D': 'diagnostyka_lab', 'G': 'genetyka', 'PS': 'patologia', 'MS': 'medycyna_sadowa',
    'C2': 'propedeutyka_chir', 'C3': 'propedeutyka_chir', 'C4': 'propedeutyka_chir',
    'C5': 'propedeutyka_chir', 'C6': 'propedeutyka_chir',
    'W': 'propedeutyka_cw', 'O': 'propedeutyka_onko', 'P': 'propedeutyka_ped',
    'PP': 'propedeutyka_psych', 'R': 'radiologia',
    # KP (propedeutyka komunikacji klinicznej) nie ma odpowiednika w istniejącym
    # rejestrze SUBJECTS - pomijamy w prototypie.
}

# Wybierz jedną grupę jako reprezentatywną (grupa 1 = wiersz 63) i wyciągnij
# sekwencję (data, kod) po kolumnach z danymi.
GROUP_ROW = 63
entries = []
for c, date in sorted(col_dates.items()):
    code = ws.cell(row=GROUP_ROW, column=c).value
    if code:
        entries.append((date, code))

# Zwiń sąsiadujące wpisy o tym samym kodzie w bloki (start, koniec, kod).
blocks = []
for date, code in entries:
    if blocks and blocks[-1][2] == code and (date - blocks[-1][1]).days <= 4:
        blocks[-1] = (blocks[-1][0], date, code)
    else:
        blocks.append((date, date, code))

print(f"Grupa {GROUP_ROW - 62} — {len(blocks)} bloków tematycznych na cały rok akademicki:\n")
out = []
ordinal_per_subject = {}
for start, end, code in blocks:
    key = CODE_TO_SUBJECT_KEY.get(code)
    label = legend.get(code, code)
    print(f"  {start} .. {end}  [{code}] {label}  -> subject_key={key}")
    if key:
        ordinal_per_subject[key] = ordinal_per_subject.get(key, 0) + 1
        out.append({
            'subject_key': key, 'ordinal': ordinal_per_subject[key],
            'title': f"{label} — blok {ordinal_per_subject[key]}",
            'starts_on': start.isoformat(), 'ends_on': end.isoformat(),
        })

with open('plan_zajec_prototype_blocks.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f"\nZapisano {len(out)} bloków do plan_zajec_prototype_blocks.json")
